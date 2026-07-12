#!/usr/bin/env python3
"""Extract the FDR parameter dictionaries, limits, and warning triggers from
the client reference workbooks into config/fdr-dictionary.json, which the
app loads at runtime.

Sources (in "Reference Files/"):
  kannada_Group_Abbreviations.xlsx  gp1 min/max limits · gp9/gp16 warning
                                    triggers · general abbreviations glossary
  kannada full form.xlsx            abbr → description per group

Re-run whenever the client sends updated dictionaries:
  python3 tools/extract_dictionaries.py "path/to/Reference Files"
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

REF_DIR = Path(sys.argv[1] if len(sys.argv) > 1 else "../Reference Files")
OUT = Path(__file__).resolve().parent.parent / "config" / "fdr-dictionary.json"

TIME_COLS = {"time", "dd", "mm", "yy", "hh", "min", "sec"}


def norm(name):
    """Normalize a parameter name for matching across files: lowercase,
    alphanumerics + '#' (Q1 engine torque and q#1 pitch rate are distinct
    parameters!), TGT→T alias (limits file says TGT451, data says T451)."""
    n = re.sub(r"[^a-z0-9#]", "", str(name).lower())
    n = re.sub(r"^tgt(\d)", r"t\1", n)
    return n


def parse_limit(val, unit):
    """Limit cells are messy: numbers, '-', 'NO LIMITS', or prose like
    '924°C (AEO) ALL ENGINES OPERATIVE' / '35 Knm'. Extract the leading
    number; scale %-unit fractions (0.9 → 90%. data is in percent)."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "-", "—") or "NO LIMIT" in s.upper():
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    num = float(m.group())
    if unit and unit.strip() == "%" and abs(num) <= 2:
        num *= 100
    return round(num, 6)


def extract():
    groups = {}

    # ---- full form: abbr → description per group ----
    wb = openpyxl.load_workbook(REF_DIR / "kannada full form.xlsx", read_only=True)
    for sn in wb.sheetnames:
        gid = "gp" + re.sub(r"\D", "", sn)
        g = groups.setdefault(gid, {"title": "", "params": {}})
        for row in wb[gid and sn].iter_rows(min_row=3, max_col=2, values_only=True):
            if not row[0] or not str(row[0]).strip():
                continue
            abbr = str(row[0]).strip()
            g["params"][norm(abbr)] = {
                "abbr": abbr,
                "description": str(row[1]).strip() if row[1] else "",
            }
    wb.close()

    # ---- abbreviations workbook: limits, triggers, glossary ----
    wb = openpyxl.load_workbook(REF_DIR / "kannada_Group_Abbreviations.xlsx", read_only=True)

    # gp1: PARAMETERS | unit | DESCRIPTION | MIN | MAX
    g1 = groups.setdefault("gp1", {"title": "", "params": {}})
    g1["title"] = "Parameter Extreme Limits"
    for row in wb["gp1"].iter_rows(min_row=4, max_col=5, values_only=True):
        if not row[0]:
            continue
        abbr = str(row[0]).strip()
        if norm(abbr) in {norm(t) for t in TIME_COLS} or abbr.upper() == "PARAMETERS":
            continue
        unit = str(row[1]).strip() if row[1] else ""
        key = norm(abbr)
        p = g1["params"].setdefault(key, {"abbr": abbr, "description": ""})
        if row[2]:
            p["description"] = str(row[2]).strip()
        p["unit"] = unit
        p["min"] = parse_limit(row[3], unit)
        p["max"] = parse_limit(row[4], unit)
        if isinstance(row[3], str) or isinstance(row[4], str):
            note = " / ".join(str(v).strip() for v in (row[3], row[4])
                              if v is not None and str(v).strip() not in ("", "-"))
            if re.search(r"[A-Za-z]{3}", note):
                p["limitNote"] = note

    # gp9 / gp16: PARAMETERS | DESCRIPTION | trigger condition
    for sn, title in (("gp9", "Systems Warnings"), ("gp16", "Engine Warnings")):
        g = groups.setdefault(sn, {"title": "", "params": {}})
        g["title"] = title
        for row in wb[sn].iter_rows(min_row=4, max_col=3, values_only=True):
            if not row[0]:
                continue
            abbr = str(row[0]).strip()
            if norm(abbr) in {norm(t) for t in TIME_COLS} or abbr.upper() == "PARAMETERS":
                continue
            key = norm(abbr)
            p = g["params"].setdefault(key, {"abbr": abbr, "description": ""})
            if row[1]:
                p["description"] = str(row[1]).strip()
            if row[2] and str(row[2]).strip():
                p["trigger"] = str(row[2]).strip()
            p["discrete"] = True

    # gp2 units (PARAMETERS | UNIT | DESCRIPTION, offset one column)
    g2 = groups.setdefault("gp2", {"title": "Navigation", "params": {}})
    for row in wb["gp2"].iter_rows(min_row=4, max_col=4, values_only=True):
        if not row[1]:
            continue
        abbr = str(row[1]).strip()
        if norm(abbr) in {norm(t) for t in TIME_COLS} or abbr.upper() == "PARAMETERS":
            continue
        key = norm(abbr)
        p = g2["params"].setdefault(key, {"abbr": abbr, "description": ""})
        if row[2]:
            p["unit"] = str(row[2]).strip()
        if row[3]:
            p["description"] = str(row[3]).strip()

    glossary = {}
    for row in wb["Abbreviations"].iter_rows(min_row=3, min_col=4, max_col=5, values_only=True):
        if row[0] and row[1]:
            glossary[str(row[0]).strip()] = str(row[1]).strip()
    wb.close()

    # discrete groups without dedicated dictionaries share the AFCS lane sets
    for gid in ("gp6", "gp10", "gp11", "gp12", "gp14", "gp15"):
        groups.setdefault(gid, {"title": "", "params": {}})
    groups["gp6"]["title"] = groups["gp10"]["title"] = "AFCS Lane 1 Status"
    groups["gp11"]["title"] = "AFCS Lane 2 Status"
    groups["gp12"]["title"] = groups["gp14"]["title"] = "AFCS Lane 1 Engagement"
    groups["gp15"]["title"] = "AFCS Lane 2 Engagement"
    groups["gp5"] = groups.get("gp5", {"title": "Air Data (ADC1)", "params": groups.get("gp5", {}).get("params", {})})
    if not groups["gp5"]["title"]:
        groups["gp5"]["title"] = "Air Data (ADC1)"

    return {"groups": groups, "glossary": glossary}


if __name__ == "__main__":
    data = extract()
    OUT.parent.mkdir(exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    nparams = sum(len(g["params"]) for g in data["groups"].values())
    nlimits = sum(1 for g in data["groups"].values() for p in g["params"].values()
                  if p.get("min") is not None or p.get("max") is not None)
    print(f"Wrote {OUT.name}: {len(data['groups'])} groups, {nparams} params, "
          f"{nlimits} with numeric limits, {len(data['glossary'])} glossary entries")
